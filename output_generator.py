import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
from config import COLORS


def create_workbook() -> openpyxl.Workbook:
    """创建新的工作簿"""
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def setup_styles():
    """设置通用样式 - 使用深色字体确保可读性"""
    styles = {
        "header_font": Font(name="微软雅黑", size=14, bold=True, color="FFFFFF"),
        "subheader_font": Font(name="微软雅黑", size=11, bold=True, color="FFFFFF"),
        "title_font": Font(name="微软雅黑", size=9, bold=True, color="FFFFFF"),
        "dark_title_font": Font(name="微软雅黑", size=9, bold=True, color="000000"),
        "normal_font": Font(name="微软雅黑", size=9, color="000000"),
        "bold_font": Font(name="微软雅黑", size=9, bold=True, color="000000"),
        "center_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left_align": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "right_align": Alignment(horizontal="right", vertical="center"),
        "header_fill": PatternFill(start_color=COLORS["header_dark"], end_color=COLORS["header_dark"], fill_type="solid"),
        "subheader_fill": PatternFill(start_color=COLORS["header_mid"], end_color=COLORS["header_mid"], fill_type="solid"),
        "parent_fill": PatternFill(start_color=COLORS["parent_green"], end_color=COLORS["parent_green"], fill_type="solid"),
        "child_fill": PatternFill(start_color=COLORS["child_light"], end_color=COLORS["child_light"], fill_type="solid"),
        "odd_fill": PatternFill(start_color=COLORS["odd_white"], end_color=COLORS["odd_white"], fill_type="solid"),
        "even_fill": PatternFill(start_color=COLORS["even_gray"], end_color=COLORS["even_gray"], fill_type="solid"),
        "independent_fill": PatternFill(start_color=COLORS["independent_orange"], end_color=COLORS["independent_orange"], fill_type="solid"),
        "summary_fill": PatternFill(start_color=COLORS["summary_yellow"], end_color=COLORS["summary_yellow"], fill_type="solid"),
        "type_fill": PatternFill(start_color=COLORS["type_title_blue"], end_color=COLORS["type_title_blue"], fill_type="solid"),
        # 项目A专用颜色（蓝色系）- 使用深蓝色配白色字体或浅蓝色配深色字体
        "project_a_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "project_a_sub_fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),
        # 项目B专用颜色（橙色系）
        "project_b_fill": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
        "project_b_sub_fill": PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid"),
        # 差额列颜色（绿色系）
        "diff_fill": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        "diff_sub_fill": PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
        # 科目列颜色
        "subject_fill": PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),
        "thin_border": Border(
            left=Side(style="thin", color=COLORS["border_gray"]),
            right=Side(style="thin", color=COLORS["border_gray"]),
            top=Side(style="thin", color=COLORS["border_gray"]),
            bottom=Side(style="thin", color=COLORS["border_gray"]),
        ),
    }
    return styles


def format_number(value, decimals=2):
    """格式化数字 - 返回字符串"""
    if value is None:
        return ""
    try:
        num = float(value)
        if decimals == 0:
            return f"{int(num):,}"
        return f"{num:,.{decimals}f}"
    except (ValueError, TypeError):
        return str(value)


def set_cell_value(cell, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    """设置单元格的值和样式"""
    if isinstance(value, str):
        cell.value = value
    elif value is None:
        cell.value = ""
    else:
        cell.value = value

    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def gen_summary_sheet(wb: openpyxl.Workbook, data: Dict[str, Any], styles: Dict):
    """生成成本对比总表"""
    ws = wb.create_sheet("成本对比总表")

    # 设置列宽（增加一列总金额差）
    col_widths = [18, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ===== Row 1: 表格标题 =====
    ws.merge_cells("A1:N1")
    set_cell_value(ws["A1"], "建安成本对比分析表",
                   font=Font(name="微软雅黑", size=16, bold=True, color="000000"),
                   fill=PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                   alignment=styles["center_align"])
    ws.row_dimensions[1].height = 35

    # ===== Row 2: 项目信息 =====
    set_cell_value(ws["A2"], "项目信息",
                   font=styles["bold_font"],
                   fill=styles["header_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])

    ws.merge_cells("B2:F2")
    set_cell_value(ws["B2"], f"{data['项目A名称']} | 建筑面积: {format_number(data['项目A总面积'], 0)}㎡ | 可售面积: {format_number(data['项目A可售面积'], 0)}㎡",
                   font=styles["bold_font"],
                   fill=styles["project_a_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(3, 7):
        ws.cell(row=2, column=col).fill = styles["project_a_fill"]
        ws.cell(row=2, column=col).border = styles["thin_border"]

    ws.merge_cells("G2:K2")
    set_cell_value(ws["G2"], f"{data['项目B名称']} | 建筑面积: {format_number(data['项目B总面积'], 0)}㎡ | 可售面积: {format_number(data['项目B可售面积'], 0)}㎡",
                   font=styles["bold_font"],
                   fill=styles["project_b_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(8, 12):
        ws.cell(row=2, column=col).fill = styles["project_b_fill"]
        ws.cell(row=2, column=col).border = styles["thin_border"]

    ws.merge_cells("L2:N2")
    set_cell_value(ws["L2"], "差额",
                   font=styles["bold_font"],
                   fill=styles["diff_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(13, 15):
        ws.cell(row=2, column=col).fill = styles["diff_fill"]
        ws.cell(row=2, column=col).border = styles["thin_border"]

    ws.row_dimensions[2].height = 28

    # ===== Row 3: 一级表头（项目名称） =====
    set_cell_value(ws["A3"], "成本科目",
                   font=styles["dark_title_font"],
                   fill=styles["subject_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])

    ws.merge_cells("B3:F3")
    set_cell_value(ws["B3"], data["项目A名称"],
                   font=styles["dark_title_font"],
                   fill=styles["project_a_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(3, 7):
        ws.cell(row=3, column=col).fill = styles["project_a_sub_fill"]
        ws.cell(row=3, column=col).border = styles["thin_border"]

    ws.merge_cells("G3:K3")
    set_cell_value(ws["G3"], data["项目B名称"],
                   font=styles["dark_title_font"],
                   fill=styles["project_b_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(8, 12):
        ws.cell(row=3, column=col).fill = styles["project_b_sub_fill"]
        ws.cell(row=3, column=col).border = styles["thin_border"]

    ws.merge_cells("L3:N3")
    set_cell_value(ws["L3"], "差额",
                   font=styles["dark_title_font"],
                   fill=styles["diff_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(13, 15):
        ws.cell(row=3, column=col).fill = styles["diff_sub_fill"]
        ws.cell(row=3, column=col).border = styles["thin_border"]

    ws.row_dimensions[3].height = 25

    # ===== Row 4: 二级表头（总价/技术经济指标） =====
    set_cell_value(ws["A4"], "",
                   fill=styles["subject_fill"],
                   border=styles["thin_border"])

    set_cell_value(ws["B4"], "总价(万元)",
                   font=styles["dark_title_font"],
                   fill=styles["project_a_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])

    ws.merge_cells("C4:D4")
    set_cell_value(ws["C4"], "相对建面指标",
                   font=styles["dark_title_font"],
                   fill=styles["project_a_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(4, 5):
        ws.cell(row=4, column=col).fill = styles["project_a_sub_fill"]
        ws.cell(row=4, column=col).border = styles["thin_border"]

    ws.merge_cells("E4:F4")
    set_cell_value(ws["E4"], "总建面指标",
                   font=styles["dark_title_font"],
                   fill=styles["project_a_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    ws.cell(row=4, column=6).fill = styles["project_a_sub_fill"]
    ws.cell(row=4, column=6).border = styles["thin_border"]

    set_cell_value(ws["G4"], "总价(万元)",
                   font=styles["dark_title_font"],
                   fill=styles["project_b_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])

    ws.merge_cells("H4:I4")
    set_cell_value(ws["H4"], "相对建面指标",
                   font=styles["dark_title_font"],
                   fill=styles["project_b_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    ws.cell(row=4, column=10).fill = styles["project_b_sub_fill"]
    ws.cell(row=4, column=10).border = styles["thin_border"]

    ws.merge_cells("J4:K4")
    set_cell_value(ws["J4"], "总建面指标",
                   font=styles["dark_title_font"],
                   fill=styles["project_b_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    ws.cell(row=4, column=11).fill = styles["project_b_sub_fill"]
    ws.cell(row=4, column=11).border = styles["thin_border"]

    set_cell_value(ws["L4"], "相对建面",
                   font=styles["dark_title_font"],
                   fill=styles["diff_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])

    set_cell_value(ws["M4"], "总建面",
                   font=styles["dark_title_font"],
                   fill=styles["diff_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])

    set_cell_value(ws["N4"], "总金额差",
                   font=styles["dark_title_font"],
                   fill=styles["diff_sub_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])

    ws.row_dimensions[4].height = 22

    # ===== Row 5: 三级表头（面积/单价） =====
    headers_row5 = [
        ("A5", "成本科目"),
        ("B5", "总价(万元)"),
        ("C5", "相对建面(㎡)"), ("D5", "相对建面单价"),
        ("E5", "总建面(㎡)"), ("F5", "总建面单价"),
        ("G5", "总价(万元)"),
        ("H5", "相对建面(㎡)"), ("I5", "相对建面单价"),
        ("J5", "总建面(㎡)"), ("K5", "总建面单价"),
        ("L5", "差额(元/㎡)"), ("M5", "差额(元/㎡)"),
        ("N5", "总金额差(万元)"),
    ]

    for cell_ref, header_text in headers_row5:
        cell = ws[cell_ref]
        col_idx = openpyxl.utils.cell.column_index_from_string(cell_ref[0])
        if col_idx >= 2 and col_idx <= 6:
            fill = styles["project_a_sub_fill"]
        elif col_idx >= 7 and col_idx <= 11:
            fill = styles["project_b_sub_fill"]
        elif col_idx >= 12:
            fill = styles["diff_sub_fill"]
        else:
            fill = styles["subject_fill"]
        set_cell_value(cell, header_text,
                       font=styles["dark_title_font"],
                       fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"])

    ws.row_dimensions[5].height = 30

    # ===== 数据行 =====
    row_idx = 6

    independent_subjects = ["前期工程", "桩基础工程", "岩土工程", "配套工程", "室外工程", "基础设施工程", "卖场包装费用"]

    for item in data["科目对比"]:
        subject = item["科目"]

        if subject in ["合计-毛坯", "总建安成本"]:
            fill = styles["summary_fill"]
            font = styles["bold_font"]
        elif subject in ["单体工程", "户内装饰工程"]:
            fill = styles["parent_fill"]
            font = styles["bold_font"]
        elif subject in independent_subjects:
            fill = styles["independent_fill"]
            font = styles["normal_font"]
        else:
            fill = styles["odd_fill"]
            font = styles["normal_font"]

        a_total_unit = item["项目A总价"] * 10000 / data["项目A总面积"] if data["项目A总面积"] > 0 else 0
        b_total_unit = item["项目B总价"] * 10000 / data["项目B总面积"] if data["项目B总面积"] > 0 else 0

        # 第一列是文本
        set_cell_value(ws.cell(row=row_idx, column=1), subject,
                       font=font, fill=fill,
                       alignment=styles["left_align"],
                       border=styles["thin_border"])

        # 数值列直接写入数值
        set_cell_value(ws.cell(row=row_idx, column=2), item["项目A总价"],
                       font=styles["normal_font"], fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"],
                       number_format='#,##0.00')

        set_cell_value(ws.cell(row=row_idx, column=3), item.get("项目A相对建面", data["项目A总面积"]),
                       font=styles["normal_font"], fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"],
                       number_format='#,##0.00')

        set_cell_value(ws.cell(row=row_idx, column=4), item["项目A建面单方"],
                       font=styles["normal_font"], fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"],
                       number_format='#,##0.00')

        set_cell_value(ws.cell(row=row_idx, column=5), data["项目A总面积"],
                       font=styles["normal_font"], fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"],
                       number_format='#,##0.00')

        # F列(6): 项目A总建面单方 = B*10000/E (公式)
        ws.cell(row=row_idx, column=6).value = f"=B{row_idx}*10000/E{row_idx}"
        ws.cell(row=row_idx, column=6).font = styles["normal_font"]
        ws.cell(row=row_idx, column=6).fill = fill
        ws.cell(row=row_idx, column=6).alignment = styles["center_align"]
        ws.cell(row=row_idx, column=6).border = styles["thin_border"]
        ws.cell(row=row_idx, column=6).number_format = '#,##0.00'

        set_cell_value(ws.cell(row=row_idx, column=7), item["项目B总价"],
                       font=styles["normal_font"], fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"],
                       number_format='#,##0.00')

        set_cell_value(ws.cell(row=row_idx, column=8), item.get("项目B相对建面", data["项目B总面积"]),
                       font=styles["normal_font"], fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"],
                       number_format='#,##0.00')

        set_cell_value(ws.cell(row=row_idx, column=9), item["项目B建面单方"],
                       font=styles["normal_font"], fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"],
                       number_format='#,##0.00')

        set_cell_value(ws.cell(row=row_idx, column=10), data["项目B总面积"],
                       font=styles["normal_font"], fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"],
                       number_format='#,##0.00')

        # K列(11): 项目B总建面单方 = G*10000/J (公式)
        ws.cell(row=row_idx, column=11).value = f"=G{row_idx}*10000/J{row_idx}"
        ws.cell(row=row_idx, column=11).font = styles["normal_font"]
        ws.cell(row=row_idx, column=11).fill = fill
        ws.cell(row=row_idx, column=11).alignment = styles["center_align"]
        ws.cell(row=row_idx, column=11).border = styles["thin_border"]
        ws.cell(row=row_idx, column=11).number_format = '#,##0.00'

        # L列(12): 相对建面单方差额 = D-I (公式)
        ws.cell(row=row_idx, column=12).value = f"=D{row_idx}-I{row_idx}"
        ws.cell(row=row_idx, column=12).font = styles["normal_font"]
        ws.cell(row=row_idx, column=12).fill = fill
        ws.cell(row=row_idx, column=12).alignment = styles["center_align"]
        ws.cell(row=row_idx, column=12).border = styles["thin_border"]
        ws.cell(row=row_idx, column=12).number_format = '#,##0.00'

        # M列(13): 总建面单方差额 = F-K (公式)
        ws.cell(row=row_idx, column=13).value = f"=F{row_idx}-K{row_idx}"
        ws.cell(row=row_idx, column=13).font = styles["normal_font"]
        ws.cell(row=row_idx, column=13).fill = fill
        ws.cell(row=row_idx, column=13).alignment = styles["center_align"]
        ws.cell(row=row_idx, column=13).border = styles["thin_border"]
        ws.cell(row=row_idx, column=13).number_format = '#,##0.00'

        # N列(14): 总金额差 = B-G (公式，项目A总价 - 项目B总价)
        ws.cell(row=row_idx, column=14).value = f"=B{row_idx}-G{row_idx}"
        ws.cell(row=row_idx, column=14).font = styles["normal_font"]
        ws.cell(row=row_idx, column=14).fill = fill
        ws.cell(row=row_idx, column=14).alignment = styles["center_align"]
        ws.cell(row=row_idx, column=14).border = styles["thin_border"]
        ws.cell(row=row_idx, column=14).number_format = '#,##0.00'

        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

        # 业态明细行
        if subject == "单体工程":
            for type_item in data.get("业态对比_单体工程", []):
                type_name_a = type_item["项目A业态"]
                type_name_b = type_item["项目B业态"]
                display_name = type_name_a if type_name_a else type_name_b

                set_cell_value(ws.cell(row=row_idx, column=1), f"  {display_name}",
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["left_align"],
                               border=styles["thin_border"])

                set_cell_value(ws.cell(row=row_idx, column=2), type_item["项目A总价"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=3), type_item.get("项目A相对建面", 0),
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=4), type_item["项目A相对建面单方"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=5), data["项目A总面积"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                # F列(6): 项目A总建面单方 = B*10000/E (公式)
                ws.cell(row=row_idx, column=6).value = f"=B{row_idx}*10000/E{row_idx}"
                ws.cell(row=row_idx, column=6).font = styles["normal_font"]
                ws.cell(row=row_idx, column=6).fill = styles["even_fill"]
                ws.cell(row=row_idx, column=6).alignment = styles["center_align"]
                ws.cell(row=row_idx, column=6).border = styles["thin_border"]
                ws.cell(row=row_idx, column=6).number_format = '#,##0.00'

                set_cell_value(ws.cell(row=row_idx, column=7), type_item["项目B总价"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=8), type_item.get("项目B相对建面", 0),
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=9), type_item["项目B相对建面单方"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=10), data["项目B总面积"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                # K列(11): 项目B总建面单方 = G*10000/J (公式)
                ws.cell(row=row_idx, column=11).value = f"=G{row_idx}*10000/J{row_idx}"
                ws.cell(row=row_idx, column=11).font = styles["normal_font"]
                ws.cell(row=row_idx, column=11).fill = styles["even_fill"]
                ws.cell(row=row_idx, column=11).alignment = styles["center_align"]
                ws.cell(row=row_idx, column=11).border = styles["thin_border"]
                ws.cell(row=row_idx, column=11).number_format = '#,##0.00'

                # L列(12): 相对建面单方差异 = D-I (公式)
                ws.cell(row=row_idx, column=12).value = f"=D{row_idx}-I{row_idx}"
                ws.cell(row=row_idx, column=12).font = styles["normal_font"]
                ws.cell(row=row_idx, column=12).fill = styles["even_fill"]
                ws.cell(row=row_idx, column=12).alignment = styles["center_align"]
                ws.cell(row=row_idx, column=12).border = styles["thin_border"]
                ws.cell(row=row_idx, column=12).number_format = '#,##0.00'

                # M列(13): 总建面单方差异 = F-K (公式)
                ws.cell(row=row_idx, column=13).value = f"=F{row_idx}-K{row_idx}"
                ws.cell(row=row_idx, column=13).font = styles["normal_font"]
                ws.cell(row=row_idx, column=13).fill = styles["even_fill"]
                ws.cell(row=row_idx, column=13).alignment = styles["center_align"]
                ws.cell(row=row_idx, column=13).border = styles["thin_border"]
                ws.cell(row=row_idx, column=13).number_format = '#,##0.00'

                # N列(14): 总金额差 = B-G (公式)
                ws.cell(row=row_idx, column=14).value = f"=B{row_idx}-G{row_idx}"
                ws.cell(row=row_idx, column=14).font = styles["normal_font"]
                ws.cell(row=row_idx, column=14).fill = styles["even_fill"]
                ws.cell(row=row_idx, column=14).alignment = styles["center_align"]
                ws.cell(row=row_idx, column=14).border = styles["thin_border"]
                ws.cell(row=row_idx, column=14).number_format = '#,##0.00'

                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

        if subject == "户内装饰工程":
            for type_item in data.get("业态对比_户内装饰工程", []):
                type_name_a = type_item["项目A业态"]
                type_name_b = type_item["项目B业态"]
                display_name = type_name_a if type_name_a else type_name_b

                set_cell_value(ws.cell(row=row_idx, column=1), f"  {display_name}",
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["left_align"],
                               border=styles["thin_border"])

                set_cell_value(ws.cell(row=row_idx, column=2), type_item["项目A总价"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=3), type_item.get("项目A相对建面", 0),
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=4), type_item["项目A相对建面单方"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=5), data["项目A总面积"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                # F列(6): 项目A总建面单方 = B*10000/E (公式)
                ws.cell(row=row_idx, column=6).value = f"=B{row_idx}*10000/E{row_idx}"
                ws.cell(row=row_idx, column=6).font = styles["normal_font"]
                ws.cell(row=row_idx, column=6).fill = styles["even_fill"]
                ws.cell(row=row_idx, column=6).alignment = styles["center_align"]
                ws.cell(row=row_idx, column=6).border = styles["thin_border"]
                ws.cell(row=row_idx, column=6).number_format = '#,##0.00'

                set_cell_value(ws.cell(row=row_idx, column=7), type_item["项目B总价"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=8), type_item.get("项目B相对建面", 0),
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=9), type_item["项目B相对建面单方"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                set_cell_value(ws.cell(row=row_idx, column=10), data["项目B总面积"],
                               font=styles["normal_font"], fill=styles["even_fill"],
                               alignment=styles["center_align"],
                               border=styles["thin_border"],
                               number_format='#,##0.00')

                # K列(11): 项目B总建面单方 = G*10000/J (公式)
                ws.cell(row=row_idx, column=11).value = f"=G{row_idx}*10000/J{row_idx}"
                ws.cell(row=row_idx, column=11).font = styles["normal_font"]
                ws.cell(row=row_idx, column=11).fill = styles["even_fill"]
                ws.cell(row=row_idx, column=11).alignment = styles["center_align"]
                ws.cell(row=row_idx, column=11).border = styles["thin_border"]
                ws.cell(row=row_idx, column=11).number_format = '#,##0.00'

                # L列(12): 相对建面单方差异 = D-I (公式)
                ws.cell(row=row_idx, column=12).value = f"=D{row_idx}-I{row_idx}"
                ws.cell(row=row_idx, column=12).font = styles["normal_font"]
                ws.cell(row=row_idx, column=12).fill = styles["even_fill"]
                ws.cell(row=row_idx, column=12).alignment = styles["center_align"]
                ws.cell(row=row_idx, column=12).border = styles["thin_border"]
                ws.cell(row=row_idx, column=12).number_format = '#,##0.00'

                # M列(13): 总建面单方差异 = F-K (公式)
                ws.cell(row=row_idx, column=13).value = f"=F{row_idx}-K{row_idx}"
                ws.cell(row=row_idx, column=13).font = styles["normal_font"]
                ws.cell(row=row_idx, column=13).fill = styles["even_fill"]
                ws.cell(row=row_idx, column=13).alignment = styles["center_align"]
                ws.cell(row=row_idx, column=13).border = styles["thin_border"]
                ws.cell(row=row_idx, column=13).number_format = '#,##0.00'

                # N列(14): 总金额差 = B-G (公式)
                ws.cell(row=row_idx, column=14).value = f"=B{row_idx}-G{row_idx}"
                ws.cell(row=row_idx, column=14).font = styles["normal_font"]
                ws.cell(row=row_idx, column=14).fill = styles["even_fill"]
                ws.cell(row=row_idx, column=14).alignment = styles["center_align"]
                ws.cell(row=row_idx, column=14).border = styles["thin_border"]
                ws.cell(row=row_idx, column=14).number_format = '#,##0.00'

                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

    return ws


def gen_fixed_sheet(wb: openpyxl.Workbook, data: Dict[str, Any], styles: Dict):
    """生成固定单方差异明细分析表"""
    ws = wb.create_sheet("固定单方差异明细分析")

    # 设置列宽
    col_widths = [25, 14, 14, 14, 14, 14, 14, 14, 14, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ===== Row 1: 大标题 =====
    ws.merge_cells("A1:J1")
    set_cell_value(ws["A1"], "固定单方差异明细分析",
                   font=Font(name="微软雅黑", size=14, bold=True, color="FFFFFF"),
                   fill=styles["header_fill"],
                   alignment=styles["center_align"])
    ws.row_dimensions[1].height = 30

    # ===== Row 2: 项目名称 =====
    # 列分配：A列=科目(1列), B-D=项目A(3列), E-G=项目B(3列), H-J=差异(3列)
    set_cell_value(ws["A2"], "成本科目",
                   font=styles["dark_title_font"],
                   fill=styles["subject_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])

    ws.merge_cells("B2:D2")
    set_cell_value(ws["B2"], f"项目A：{data['项目A名称']}",
                   font=styles["dark_title_font"],
                   fill=styles["project_a_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(3, 5):
        ws.cell(row=2, column=col).fill = styles["project_a_fill"]
        ws.cell(row=2, column=col).border = styles["thin_border"]

    ws.merge_cells("E2:G2")
    set_cell_value(ws["E2"], f"项目B：{data['项目B名称']}",
                   font=styles["dark_title_font"],
                   fill=styles["project_b_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(6, 8):
        ws.cell(row=2, column=col).fill = styles["project_b_fill"]
        ws.cell(row=2, column=col).border = styles["thin_border"]

    ws.merge_cells("H2:J2")
    set_cell_value(ws["H2"], "差异分析",
                   font=styles["dark_title_font"],
                   fill=styles["diff_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(9, 11):
        ws.cell(row=2, column=col).fill = styles["diff_fill"]
        ws.cell(row=2, column=col).border = styles["thin_border"]

    ws.row_dimensions[2].height = 25

    # ===== Row 3: 列标题 =====
    headers = [
        "科目/明细",
        "项目A-单方", "项目A-含量", "项目A-单价(元)",
        "项目B-单方", "项目B-含量", "项目B-单价(元)",
        "单方差异", "价格差异影响", "含量差异影响"
    ]

    for i, header in enumerate(headers, 1):
        col_idx = i
        if col_idx >= 2 and col_idx <= 4:
            fill = styles["project_a_sub_fill"]
        elif col_idx >= 5 and col_idx <= 7:
            fill = styles["project_b_sub_fill"]
        elif col_idx >= 8:
            fill = styles["diff_sub_fill"]
        else:
            fill = styles["subject_fill"]
        set_cell_value(ws.cell(row=3, column=i), header,
                       font=styles["dark_title_font"],
                       fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"])

    ws.row_dimensions[3].height = 35

    # ===== 数据行 =====
    row_idx = 4
    for type_analysis in data["业态分析"]:
        # 业态标题行
        ws.merge_cells(f"A{row_idx}:J{row_idx}")
        set_cell_value(ws.cell(row=row_idx, column=1),
                       f"{type_analysis['业态A']} VS {type_analysis['业态B']}",
                       font=styles["bold_font"],
                       fill=styles["type_fill"],
                       alignment=styles["center_align"],
                       border=styles["thin_border"])
        for col in range(2, 11):
            ws.cell(row=row_idx, column=col).fill = styles["type_fill"]
            ws.cell(row=row_idx, column=col).border = styles["thin_border"]
        ws.row_dimensions[row_idx].height = 25
        row_idx += 1

        # 初始化累加变量
        total_a_unit = 0
        total_b_unit = 0
        total_diff = 0
        total_price_impact = 0
        total_content_impact = 0

        for j, item in enumerate(type_analysis["科目明细"]):
            is_construction_child = item["科目编码"].startswith("03.04.01") and len(item["科目编码"].split(".")) == 4
            is_construction_parent = item["科目编码"] == "03.04.01"

            if is_construction_parent:
                fill = styles["parent_fill"]
                font = styles["bold_font"]
                indent = ""
            elif is_construction_child:
                fill = styles["child_fill"]
                font = styles["normal_font"]
                indent = "  "
            else:
                fill = styles["odd_fill"] if j % 2 == 0 else styles["even_fill"]
                font = styles["normal_font"]
                indent = ""

            # 累加所有非四级子科目的数据（建筑工程的四级子项不单独累加，避免重复计算）
            if not is_construction_child:
                total_a_unit += item["项目A单方"]
                total_b_unit += item["项目B单方"]
                total_diff += (item["项目A单方"] - item["项目B单方"])
                total_price_impact += item["价格差异影响"]
                total_content_impact += item["含量差异影响"]

            # 科目名称列
            set_cell_value(ws.cell(row=row_idx, column=1), indent + item["科目名称"],
                           font=font, fill=fill,
                           alignment=styles["left_align"],
                           border=styles["thin_border"])

            # 项目A数据
            set_cell_value(ws.cell(row=row_idx, column=2), item["项目A单方"],
                           font=font, fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            set_cell_value(ws.cell(row=row_idx, column=3), item["项目A含量"],
                           font=font, fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            set_cell_value(ws.cell(row=row_idx, column=4), item["项目A单价"],
                           font=font, fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            # 项目B数据
            set_cell_value(ws.cell(row=row_idx, column=5), item["项目B单方"],
                           font=font, fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            set_cell_value(ws.cell(row=row_idx, column=6), item["项目B含量"],
                           font=font, fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            set_cell_value(ws.cell(row=row_idx, column=7), item["项目B单价"],
                           font=font, fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            # H列(8): 单方差异 = B-E (公式)
            ws.cell(row=row_idx, column=8).value = f"=B{row_idx}-E{row_idx}"
            ws.cell(row=row_idx, column=8).font = font
            ws.cell(row=row_idx, column=8).fill = fill
            ws.cell(row=row_idx, column=8).alignment = styles["center_align"]
            ws.cell(row=row_idx, column=8).border = styles["thin_border"]
            ws.cell(row=row_idx, column=8).number_format = '#,##0.00'

            # I列(9): 价格差异影响 = (D-G)*C (公式)
            ws.cell(row=row_idx, column=9).value = f"=(D{row_idx}-G{row_idx})*C{row_idx}"
            ws.cell(row=row_idx, column=9).font = font
            ws.cell(row=row_idx, column=9).fill = fill
            ws.cell(row=row_idx, column=9).alignment = styles["center_align"]
            ws.cell(row=row_idx, column=9).border = styles["thin_border"]
            ws.cell(row=row_idx, column=9).number_format = '#,##0.00'

            # J列(10): 含量差异影响 = G*(C-F) (公式)
            ws.cell(row=row_idx, column=10).value = f"=G{row_idx}*(C{row_idx}-F{row_idx})"
            ws.cell(row=row_idx, column=10).font = font
            ws.cell(row=row_idx, column=10).fill = fill
            ws.cell(row=row_idx, column=10).alignment = styles["center_align"]
            ws.cell(row=row_idx, column=10).border = styles["thin_border"]
            ws.cell(row=row_idx, column=10).number_format = '#,##0.00'

            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

        # 合计行
        ws.merge_cells(f"A{row_idx}:A{row_idx}")
        set_cell_value(ws.cell(row=row_idx, column=1), "固定成本合计",
                       font=styles["bold_font"],
                       fill=styles["summary_fill"],
                       alignment=styles["left_align"],
                       border=styles["thin_border"])

        summary_cols = [
            (2, total_a_unit, '#,##0.00'),
            (3, "", None),
            (4, "", None),
            (5, total_b_unit, '#,##0.00'),
            (6, "", None),
            (7, "", None),
            (8, total_diff, '#,##0.00'),
            (9, total_price_impact, '#,##0.00'),
            (10, total_content_impact, '#,##0.00'),
        ]

        for col, value, num_format in summary_cols:
            set_cell_value(ws.cell(row=row_idx, column=col), value,
                           font=styles["bold_font"],
                           fill=styles["summary_fill"],
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format=num_format)

        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

    return ws


def gen_elastic_sheet(wb: openpyxl.Workbook, data: Dict[str, Any], styles: Dict):
    """生成弹性单方差异明细分析表"""
    ws = wb.create_sheet("弹性单方差异明细分析")

    # 设置列宽
    col_widths = [25, 14, 16, 16, 14, 16, 16, 14, 35]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ===== Row 1: 大标题 =====
    ws.merge_cells("A1:I1")
    set_cell_value(ws["A1"], "弹性单方差异明细分析",
                   font=Font(name="微软雅黑", size=14, bold=True, color="FFFFFF"),
                   fill=styles["header_fill"],
                   alignment=styles["center_align"])
    ws.row_dimensions[1].height = 30

    # ===== Row 2: 项目名称 =====
    # 列分配：A列=科目(1列), B-D=项目A(3列), E-G=项目B(3列), H-I=差异(2列)
    set_cell_value(ws["A2"], "成本科目",
                   font=styles["dark_title_font"],
                   fill=styles["subject_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])

    ws.merge_cells("B2:D2")
    set_cell_value(ws["B2"], f"项目A：{data['项目A名称']}",
                   font=styles["dark_title_font"],
                   fill=styles["project_a_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(3, 5):
        ws.cell(row=2, column=col).fill = styles["project_a_fill"]
        ws.cell(row=2, column=col).border = styles["thin_border"]

    ws.merge_cells("E2:G2")
    set_cell_value(ws["E2"], f"项目B：{data['项目B名称']}",
                   font=styles["dark_title_font"],
                   fill=styles["project_b_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    for col in range(6, 8):
        ws.cell(row=2, column=col).fill = styles["project_b_fill"]
        ws.cell(row=2, column=col).border = styles["thin_border"]

    ws.merge_cells("H2:I2")
    set_cell_value(ws["H2"], "差异分析",
                   font=styles["dark_title_font"],
                   fill=styles["diff_fill"],
                   alignment=styles["center_align"],
                   border=styles["thin_border"])
    ws.cell(row=2, column=9).fill = styles["diff_fill"]
    ws.cell(row=2, column=9).border = styles["thin_border"]

    ws.row_dimensions[2].height = 25

    # ===== Row 3: 列标题 =====
    headers = [
        "科目/明细",
        "项目A-单方", "项目A-含量指标", "项目A-配置单价",
        "项目B-单方", "项目B-含量指标", "项目B-配置单价",
        "单方差异", "差异说明"
    ]

    for i, header in enumerate(headers, 1):
        col_idx = i
        if col_idx >= 2 and col_idx <= 4:
            fill = styles["project_a_sub_fill"]
        elif col_idx >= 5 and col_idx <= 7:
            fill = styles["project_b_sub_fill"]
        elif col_idx >= 8:
            fill = styles["diff_sub_fill"]
        else:
            fill = styles["subject_fill"]
        set_cell_value(ws.cell(row=3, column=i), header,
                       font=styles["dark_title_font"],
                       fill=fill,
                       alignment=styles["center_align"],
                       border=styles["thin_border"])

    ws.row_dimensions[3].height = 35

    # ===== 数据行 =====
    row_idx = 4
    for type_analysis in data["业态分析"]:
        # 业态标题行
        ws.merge_cells(f"A{row_idx}:I{row_idx}")
        set_cell_value(ws.cell(row=row_idx, column=1),
                       f"{type_analysis['业态A']} VS {type_analysis['业态B']}",
                       font=styles["bold_font"],
                       fill=styles["type_fill"],
                       alignment=styles["center_align"],
                       border=styles["thin_border"])
        for col in range(2, 10):
            ws.cell(row=row_idx, column=col).fill = styles["type_fill"]
            ws.cell(row=row_idx, column=col).border = styles["thin_border"]
        ws.row_dimensions[row_idx].height = 25
        row_idx += 1

        total_a_unit = 0
        total_b_unit = 0
        total_diff = 0

        for j, item in enumerate(type_analysis["科目明细"]):
            fill = styles["odd_fill"] if j % 2 == 0 else styles["even_fill"]

            diff_desc = item.get("差异说明", "")
            if not diff_desc:
                if item.get("含量差异影响") or item.get("配置差异影响"):
                    parts = []
                    if item.get("含量差异影响") and abs(item["含量差异影响"]) > 0.01:
                        parts.append(f"含量影响{format_number(item['含量差异影响'])}元/㎡")
                    if item.get("配置差异影响") and abs(item["配置差异影响"]) > 0.01:
                        parts.append(f"配置影响{format_number(item['配置差异影响'])}元/㎡")
                    diff_desc = "；".join(parts)

            total_a_unit += item["项目A单方"]
            total_b_unit += item["项目B单方"]
            total_diff += item["单方差异"]

            # 科目名称
            set_cell_value(ws.cell(row=row_idx, column=1), item["科目名称"],
                           font=styles["normal_font"], fill=fill,
                           alignment=styles["left_align"],
                           border=styles["thin_border"])

            # 项目A数据
            set_cell_value(ws.cell(row=row_idx, column=2), item["项目A单方"],
                           font=styles["normal_font"], fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            set_cell_value(ws.cell(row=row_idx, column=3), item.get("项目A含量指标", 0),
                           font=styles["normal_font"], fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            set_cell_value(ws.cell(row=row_idx, column=4), item.get("项目A配置单价", 0),
                           font=styles["normal_font"], fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            # 项目B数据
            set_cell_value(ws.cell(row=row_idx, column=5), item["项目B单方"],
                           font=styles["normal_font"], fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            set_cell_value(ws.cell(row=row_idx, column=6), item.get("项目B含量指标", 0),
                           font=styles["normal_font"], fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            set_cell_value(ws.cell(row=row_idx, column=7), item.get("项目B配置单价", 0),
                           font=styles["normal_font"], fill=fill,
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format='#,##0.00')

            # H列(8): 单方差异 = B-E (公式)
            ws.cell(row=row_idx, column=8).value = f"=B{row_idx}-E{row_idx}"
            ws.cell(row=row_idx, column=8).font = styles["normal_font"]
            ws.cell(row=row_idx, column=8).fill = fill
            ws.cell(row=row_idx, column=8).alignment = styles["center_align"]
            ws.cell(row=row_idx, column=8).border = styles["thin_border"]
            ws.cell(row=row_idx, column=8).number_format = '#,##0.00'

            # I列(9): 差异说明 (文本)
            set_cell_value(ws.cell(row=row_idx, column=9), diff_desc,
                           font=styles["normal_font"], fill=fill,
                           alignment=styles["left_align"],
                           border=styles["thin_border"])

            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

        # 合计行
        ws.merge_cells(f"A{row_idx}:A{row_idx}")
        set_cell_value(ws.cell(row=row_idx, column=1), "弹性成本合计",
                       font=styles["bold_font"],
                       fill=styles["summary_fill"],
                       alignment=styles["left_align"],
                       border=styles["thin_border"])

        summary_cols = [
            (2, total_a_unit, '#,##0.00'),
            (3, "", None),
            (4, "", None),
            (5, total_b_unit, '#,##0.00'),
            (6, "", None),
            (7, "", None),
            (8, total_diff, '#,##0.00'),
            (9, "", None),
        ]

        for col, value, num_format in summary_cols:
            set_cell_value(ws.cell(row=row_idx, column=col), value,
                           font=styles["bold_font"],
                           fill=styles["summary_fill"],
                           alignment=styles["center_align"],
                           border=styles["thin_border"],
                           number_format=num_format)

        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

    return ws


def generate_output(summary_data: Dict, fixed_data: Dict, elastic_data: Dict, output_path: str):
    """生成最终的Excel输出文件"""
    wb = create_workbook()
    styles = setup_styles()

    gen_summary_sheet(wb, summary_data, styles)
    gen_fixed_sheet(wb, fixed_data, styles)
    gen_elastic_sheet(wb, elastic_data, styles)

    wb.save(output_path)
    wb.close()

    print(f"输出文件已保存: {output_path}")
