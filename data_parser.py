import openpyxl
from typing import Dict, List, Any, Optional


def safe_float(value):
    """安全地将值转换为float，处理非数字内容"""
    if value is None:
        return 0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0


def find_sheet(wb, possible_names):
    """在工作簿中查找可能的工作表"""
    for name in possible_names:
        if name in wb.sheetnames:
            return wb[name]
    # 如果都找不到，尝试模糊匹配
    for sheet_name in wb.sheetnames:
        for name in possible_names:
            if name in sheet_name or sheet_name in name:
                return wb[sheet_name]
    return None


def parse_basic_info(wb: openpyxl.Workbook) -> Dict[str, Any]:
    """解析02项目基础信息表（如果不存在则返回空字典）"""
    # 支持多种可能的工作表名称
    possible_names = ["02 项目基础信息（输入）", "项目基础信息", "02项目基础信息", "基础信息"]
    ws = find_sheet(wb, possible_names)
    if ws is None:
        # 如果02表不存在，返回空字典，后续从其他表获取信息
        return {}
    
    info = {}
    
    # 读取建筑面积信息（根据实际格式调整）
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] and "建筑面积" in str(row[0]):
            info["总建筑面积"] = row[1] if len(row) > 1 else None
        if row[0] and "可售面积" in str(row[0]):
            info["可售面积"] = row[1] if len(row) > 1 else None
    
    return info


def parse_cost_control(wb: openpyxl.Workbook, extracted_name: str = "") -> Dict[str, Any]:
    """解析04目标成本控制表（支持多种格式）"""
    # 支持多种可能的工作表名称
    possible_names = ["04 目标成本控制表", "目标成本控制表", "04目标成本控制表", "成本控制表"]
    ws = find_sheet(wb, possible_names)
    if ws is None:
        raise ValueError("未找到目标成本控制表，请确保Excel文件包含名为'04 目标成本控制表'或类似名称的工作表")

    result = {
        "项目信息": {},
        "科目汇总": {},
        "业态数据": {},
        "汇总行": {},
    }

    # 检测数据单位（万元或元）
    is_in_wan_yuan = True
    for header_row_idx in [4, 5, 6]:
        header_row = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
        for cell in header_row:
            if cell and isinstance(cell, str):
                if "万元" in cell:
                    is_in_wan_yuan = True
                    break
                elif "万元" not in cell and ("元" in cell or "总价" in str(cell)):
                    if "（元）" in cell or "(元)" in cell or "_x000d_\n(元)" in cell:
                        is_in_wan_yuan = False
                        break
        if not is_in_wan_yuan:
            break
    
    unit_factor = 1.0 if is_in_wan_yuan else 0.0001

    # 尝试从第3行获取项目信息（新格式）
    row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    project_name = ""
    total_area = 0
    saleable_area = 0
    
    # 检查是否是新格式（第3行包含"项目名称："）
    if len(row3) > 0 and row3[0] and "项目名称" in str(row3[0]):
        # 新格式：第3行包含项目名称和建筑面积
        if len(row3) > 2 and row3[2]:
            project_name = str(row3[2]).strip()
            project_name = project_name.replace('_x000d_\n', '').replace('\n', '')
        
        # 查找建筑面积信息
        for i, val in enumerate(row3):
            if val:
                # 处理标签在当前单元格的情况
                if isinstance(val, str) and "建筑面积" in val:
                    if i + 1 < len(row3) and row3[i + 1]:
                        total_area = row3[i + 1]
                elif isinstance(val, str) and ("可租售面积" in val or "可售面积" in val):
                    if i + 1 < len(row3) and row3[i + 1]:
                        saleable_area = row3[i + 1]
                # 处理值在当前单元格的情况，检查前后单元格是否有标签
                elif isinstance(val, (int, float)) and val > 0:
                    if i > 0 and row3[i - 1]:
                        label = str(row3[i - 1])
                        if "建筑面积" in label:
                            total_area = val
                        elif "可租售面积" in label or "可售面积" in label:
                            saleable_area = val
                    if i + 1 < len(row3) and isinstance(row3[i + 1], str):
                        label = str(row3[i + 1])
                        if "建筑面积" in label:
                            total_area = val
                        elif "可租售面积" in label or "可售面积" in label:
                            saleable_area = val
    
    # 如果第3行不是新格式，尝试从第4行获取（旧格式）
    if not project_name:
        row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        if len(row4) > 2 and row4[2]:
            project_name = str(row4[2]).strip()
            project_name = project_name.replace('_x000d_\n', '').replace('\n', '')
        elif len(row4) > 1 and row4[1]:
            project_name = str(row4[1]).strip()
            project_name = project_name.replace('_x000d_\n', '').replace('\n', '')
        
        if total_area == 0:
            total_area = row4[5] if len(row4) > 5 and row4[5] else 0
        if saleable_area == 0:
            saleable_area = row4[9] if len(row4) > 9 and row4[9] else 0

    # 如果还是空，尝试从02表获取
    if not project_name:
        try:
            ws02 = wb["02 项目基础信息（输入）"]
            for row in ws02.iter_rows(min_row=1, max_row=20, values_only=True):
                if row[0] and "项目名称" in str(row[0]):
                    project_name = str(row[3]) if len(row) > 3 and row[3] else ""
                    project_name = project_name.replace('_x000d_\n', '').replace('\n', '')
                    break
        except:
            pass

    template_names = ["集团商业标准成本", "苏州公司U3+档毛坯标准刻度", "标准", "模板"]
    is_template_name = any(template in project_name for template in template_names)
    if is_template_name and extracted_name:
        project_name = extracted_name
    
    if not project_name and extracted_name:
        project_name = extracted_name

    result["项目信息"]["名称"] = project_name

    total_area = safe_float(total_area)
    saleable_area = safe_float(saleable_area)
    
    if total_area == 0 or saleable_area == 0:
        try:
            info_from_05 = get_project_info_from_sheet(wb, "05 成本测算明细")
            if total_area == 0 and info_from_05.get("总建筑面积"):
                total_area = safe_float(info_from_05["总建筑面积"])
            if saleable_area == 0 and info_from_05.get("可售面积"):
                saleable_area = safe_float(info_from_05["可售面积"])
            if saleable_area == 0 and info_from_05.get("可租售面积"):
                saleable_area = safe_float(info_from_05["可租售面积"])
        except:
            pass
    
    if total_area == 0 or saleable_area == 0:
        try:
            info_from_02 = get_project_info_from_sheet(wb, "02 项目基础信息（输入）")
            if total_area == 0 and info_from_02.get("总建筑面积"):
                total_area = safe_float(info_from_02["总建筑面积"])
            if saleable_area == 0 and info_from_02.get("可售面积"):
                saleable_area = safe_float(info_from_02["可售面积"])
            if saleable_area == 0 and info_from_02.get("可租售面积"):
                saleable_area = safe_float(info_from_02["可租售面积"])
        except:
            pass
    
    result["项目信息"]["总建筑面积"] = total_area
    result["项目信息"]["可售面积"] = saleable_area
    
    current_subject = None
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        a_val = row[0] if row[0] else None
        b_val = row[1] if row[1] else None
        
        if not a_val and not b_val:
            continue
        
        if a_val and a_val in ["前期工程", "桩基础工程", "岩土工程", "单体工程", 
                                "配套工程", "室外工程", "户内装饰工程", 
                                "基础设施工程", "卖场包装费用", "其它"]:
            current_subject = a_val
            result["科目汇总"][current_subject] = {
                "总价": safe_float(row[2]) * unit_factor if len(row) > 2 and row[2] else 0,
                "相对建面": row[3] if len(row) > 3 and row[3] else 0,
                "相对建面单方": row[4] if len(row) > 4 and row[4] else 0,
                "相对建造面积": row[5] if len(row) > 5 and row[5] else 0,
                "相对建造单方": row[6] if len(row) > 6 and row[6] else 0,
                "可售面积": row[7] if len(row) > 7 and row[7] else 0,
                "可售单方": row[8] if len(row) > 8 and row[8] else 0,
            }
        
        elif a_val and a_val in ["合计", "总建安成本"]:
            key = "合计-毛坯" if a_val == "合计" else "总建安成本"
            result["汇总行"][key] = {
                "总价": safe_float(row[2]) * unit_factor if len(row) > 2 and row[2] else 0,
                "相对建面": row[3] if len(row) > 3 and row[3] else 0,
                "相对建面单方": row[4] if len(row) > 4 and row[4] else 0,
                "相对建造面积": row[5] if len(row) > 5 and row[5] else 0,
                "相对建造单方": row[6] if len(row) > 6 and row[6] else 0,
                "可售面积": row[7] if len(row) > 7 and row[7] else 0,
                "可售单方": row[8] if len(row) > 8 and row[8] else 0,
            }
        
        elif b_val and not b_val.startswith("$") and current_subject in ["单体工程", "户内装饰工程"]:
            total_price = safe_float(row[2]) * unit_factor if len(row) > 2 and row[2] else 0
            area = row[3] if len(row) > 3 and row[3] else 0
            
            if total_price > 0 or area > 0:
                type_name = b_val
                if "(住宅)" in type_name:
                    type_name = type_name.replace("(住宅)", "")
                if "(商业)" in type_name:
                    type_name = type_name.replace("(商业)", "")
                
                if "叠墅" in type_name or "叠加" in type_name:
                    type_name = "叠墅"
                
                if type_name in result["业态数据"]:
                    existing_data = result["业态数据"][type_name]
                    if current_subject not in existing_data:
                        result["业态数据"][type_name][current_subject] = {
                            "总价": total_price,
                            "相对建面": area,
                            "相对建面单方": row[4] if len(row) > 4 and row[4] else 0,
                            "相对建造面积": row[5] if len(row) > 5 and row[5] else 0,
                            "相对建造单方": row[6] if len(row) > 6 and row[6] else 0,
                            "可售面积": row[7] if len(row) > 7 and row[7] else 0,
                            "可售单方": row[8] if len(row) > 8 and row[8] else 0,
                        }
                else:
                    matched_type = None
                    for existing_type, existing_data in result["业态数据"].items():
                        if existing_type != type_name:
                            existing_area = 0
                            for subject_data in existing_data.values():
                                existing_area = subject_data.get("相对建面", 0)
                                if existing_area > 0:
                                    break
                            if abs(existing_area - area) < 0.01 and current_subject not in existing_data:
                                matched_type = existing_type
                                break
                    
                    if matched_type:
                        result["业态数据"][matched_type][current_subject] = {
                            "总价": total_price,
                            "相对建面": area,
                            "相对建面单方": row[4] if len(row) > 4 and row[4] else 0,
                            "相对建造面积": row[5] if len(row) > 5 and row[5] else 0,
                            "相对建造单方": row[6] if len(row) > 6 and row[6] else 0,
                            "可售面积": row[7] if len(row) > 7 and row[7] else 0,
                            "可售单方": row[8] if len(row) > 8 and row[8] else 0,
                        }
                    else:
                        result["业态数据"][type_name] = {}
                        result["业态数据"][type_name][current_subject] = {
                            "总价": total_price,
                            "相对建面": area,
                            "相对建面单方": row[4] if len(row) > 4 and row[4] else 0,
                            "相对建造面积": row[5] if len(row) > 5 and row[5] else 0,
                            "相对建造单方": row[6] if len(row) > 6 and row[6] else 0,
                            "可售面积": row[7] if len(row) > 7 and row[7] else 0,
                            "可售单方": row[8] if len(row) > 8 and row[8] else 0,
                        }
        
        elif current_subject == "其它" and b_val and not b_val.startswith("$"):
            subject_name = b_val
            result["科目汇总"][subject_name] = {
                "总价": safe_float(row[2]) * unit_factor if len(row) > 2 and row[2] else 0,
                "相对建面": row[3] if len(row) > 3 and row[3] else 0,
                "相对建面单方": row[4] if len(row) > 4 and row[4] else 0,
                "相对建造面积": row[5] if len(row) > 5 and row[5] else 0,
                "相对建造单方": row[6] if len(row) > 6 and row[6] else 0,
                "可售面积": row[7] if len(row) > 7 and row[7] else 0,
                "可售单方": row[8] if len(row) > 8 and row[8] else 0,
            }
    
    return result


def get_project_info_from_sheet(ws, sheet_name):
    """从指定工作表获取项目信息"""
    try:
        ws_target = ws[sheet_name]
        info = {}
        for row in ws_target.iter_rows(min_row=1, max_row=30, values_only=True):
            if row[0] and isinstance(row[0], str):
                if "建筑面积" in row[0]:
                    info["总建筑面积"] = row[1] if len(row) > 1 else None
                elif "可售面积" in row[0] or "可租售面积" in row[0]:
                    info["可售面积"] = row[1] if len(row) > 1 else None
        return info
    except:
        return {}


# 弹性科目列表（当没有属性列时使用）
ELASTIC_SUBJECTS = [
    "入户门工程", "外立面门窗工程", "外立面装饰工程", 
    "栏杆工程", "公共部位装饰工程", "安防系统", "电梯工程"
]


def load_project(file_path: str) -> Dict[str, Any]:
    """加载项目数据"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 从文件名中提取项目名称（去掉扩展名和路径）
    import os
    extracted_name = os.path.basename(file_path)
    extracted_name = os.path.splitext(extracted_name)[0]
    
    return {
        "成本控制": parse_cost_control(wb, extracted_name),
        "测算明细": parse_cost_detail(wb),
        "基础信息": parse_basic_info(wb),
    }


def get_project_types(project_data: Dict) -> List[str]:
    """从项目数据中获取业态列表"""
    types = []
    
    # 从成本控制数据获取业态
    if "成本控制" in project_data and "业态数据" in project_data["成本控制"]:
        types.extend(project_data["成本控制"]["业态数据"].keys())
    
    # 从测算明细获取业态（作为补充）
    if "测算明细" in project_data and "业态明细" in project_data["测算明细"]:
        for type_name in project_data["测算明细"]["业态明细"].keys():
            if type_name not in types:
                types.append(type_name)
    
    return sorted(types)


def parse_cost_detail(wb: openpyxl.Workbook) -> Dict[str, Any]:
    """解析05成本测算明细表
    
    支持两种格式：
    格式A（摩天轮）：
    - 属性列：S列（索引18），值为"固定"、"弹性"或"暂定"
    - G列(6):含量/系数值；I列(8):工程量；K列(10):单价；O列(14):相对建筑面积单方
    
    格式B（月陇云岚）：
    - F列(5):含量值；H列(7):工程量；J列(9):综合单价；L列(11):单方（相对建筑面积）
    """
    possible_names = ["05 成本测算明细", "成本测算明细", "05成本测算明细", "测算明细"]
    ws = find_sheet(wb, possible_names)
    if ws is None:
        raise ValueError("未找到成本测算明细表")
    
    result = {
        "科目属性": {},
        "业态明细": {},
    }
    
    current_type = None
    
    type_keywords = [
        "地下室其他(住宅)", "地下室其他", "普通多层", "小高层", "高层", 
        "洋房", "叠拼", "叠墅", "联排", "其他公共配套", "幼儿园",
        "地库及机房(商业)", "集中商业", "写字楼", "酒店", "地库及机房", "叠加别墅"
    ]
    
    # 检测数据格式
    header_row = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    is_format_b = False
    if len(header_row) >= 2:
        row1 = header_row[0]
        # 检查第10列（索引9）是否包含"综合单价"，这是格式B（月陇云岚）的特征
        if len(row1) > 9 and row1[9] and "综合单价" in str(row1[9]):
            is_format_b = True
    
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        code_name = row[0] if row[0] else None
        if not code_name:
            continue
        
        code_name_str = str(code_name).strip()
        
        # 识别业态名称行
        is_type_row = False
        for keyword in type_keywords:
            if code_name_str == keyword or keyword in code_name_str:
                is_type_row = True
                break
        
        if not is_type_row and not code_name_str.startswith("03.") and not code_name_str.startswith(" "):
            if len(code_name_str) < 30 and any(k in code_name_str for k in ["地下室", "小高层", "叠墅", "洋房", "公共配套", "多层", "高层", "商业", "办公", "地库", "叠加"]):
                is_type_row = True
        
        if is_type_row:
            current_type = code_name_str.replace("(住宅)", "").replace("(商业)", "")
            
            if current_type not in result["业态明细"]:
                result["业态明细"][current_type] = {}
            continue

        # 识别科目行（以03.04.开头的行，可能有缩进）
        if current_type and (code_name_str.startswith("03.04.") or code_name_str.startswith("    03.04.")):
            clean_code = code_name_str.strip()
            
            # 分割编码和名称
            parts = clean_code.split(None, 1)
            if len(parts) >= 2:
                code = parts[0].strip()
                name = parts[1].strip()
            else:
                code = ""
                name = ""
                for i, char in enumerate(clean_code):
                    if '\u4e00' <= char <= '\u9fff':
                        code = clean_code[:i].strip()
                        name = clean_code[i:].strip()
                        break
                if not code:
                    continue
            
            # 根据格式读取数据
            if is_format_b:
                # 格式B（月陇云岚）
                # 列结构：A-科目名(0), B-说明(1), C-建筑面积(2), D-单位(3), E-含量说明(4), F-含量值(5), G-含量单位(6), H-工程量(7), I-工程量单位(8), J-综合单价(9), K-单价单位(10), L-总价(11), M-相对建筑面积(12), N-相对建筑面积单方(13), O-相对建造面积(14)
                hanliang_val = row[5] if len(row) > 5 and row[5] else 0  # F列-含量值
                hanliang_unit = row[6] if len(row) > 6 and row[6] else ""  # G列-含量单位
                gongchengliang_val = row[7] if len(row) > 7 and row[7] else 0  # H列-工程量
                danjia_val = row[9] if len(row) > 9 and row[9] else 0  # J列-综合单价
                danfang_val = row[13] if len(row) > 13 and row[13] else 0  # N列-相对建筑面积单方
                
                # 如果N列没有数据，尝试通过总价/工程量计算单方
                if danfang_val == 0:
                    hanliang_float = safe_float(hanliang_val)
                    danjia_float = safe_float(danjia_val)
                    if hanliang_float > 0 and danjia_float > 0:
                        danfang_val = hanliang_float * danjia_float  # 含量×单价
                    else:
                        gongchengliang_float = safe_float(gongchengliang_val)
                        total_price = row[11] if len(row) > 11 and row[11] else 0  # L列-总价
                        if gongchengliang_float > 0 and total_price > 0:
                            # 总价转换为万元，工程量单位据说是"元"，需要转换
                            danfang_val = safe_float(total_price) * 10000 / safe_float(gongchengliang_float) if gongchengliang_float > 0 else 0
                
                # 格式B没有属性列，根据科目名称判断
                attr = None
                if code.count(".") == 2:  # 三级科目
                    if any(es in name for es in ELASTIC_SUBJECTS):
                        attr = "弹性"
                    else:
                        attr = "固定"
            else:
                # 格式A（摩天轮）
                danfang_val = row[14] if len(row) > 14 and row[14] else 0  # O列-单方
                hanliang_val = row[6] if len(row) > 6 and row[6] else 0  # G列-含量
                hanliang_unit = ""  # 格式A没有单独的含量单位列
                gongchengliang_val = row[8] if len(row) > 8 and row[8] else 0  # I列-工程量
                danjia_val = row[10] if len(row) > 10 and row[10] else 0  # K列-单价
                attr = row[18] if len(row) > 18 and row[18] else None  # S列-属性
            
            # 记录属性
            if attr and attr in ["固定", "弹性", "暂定"]:
                if code.count(".") == 2:
                    result["科目属性"][code] = attr
            
            # 跳过空数据行
            existing = result["业态明细"][current_type].get(code)
            if existing and safe_float(danfang_val) == 0 and safe_float(danjia_val) == 0 and safe_float(hanliang_val) == 0:
                continue

            data = {
                "编码": code,
                "名称": name,
                "单方": danfang_val,
                "含量": hanliang_val,
                "含量单位": hanliang_unit,
                "工程量": gongchengliang_val,
                "含量说明": row[4] if len(row) > 4 and row[4] else "",
                "单价": danjia_val,
                "属性": attr,
            }
            
            result["业态明细"][current_type][code] = data
    
    return result